
from flask import Flask, render_template, request, send_file, redirect, Response
from utils import load_history, save_history
from markupsafe import Markup
from io import BytesIO, StringIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import re
import csv
import json
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Table,
    TableStyle,
    Spacer,
    KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.textlabels import Label

app = Flask(__name__)

latest_results_text = ""
latest_results_rows = []

HISTORY_FILE = "history.json"
history = load_history()
latest_filtered_history = []

def parse_log_datetime(line):
    try:
        text = line[:19]
        return datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None 

def highlight_keyword(line, keyword, case_sensitive=False):
    if not keyword:
        return line

    flags = 0 if case_sensitive else re.IGNORECASE

    pattern = re.compile(re.escape(keyword), flags)

    highlighted = pattern.sub(
        lambda match: f"<mark>{match.group(0)}</mark>",
        line
    )

    return Markup(highlighted)

def clean_export_results(results):
    cleaned = []

    for line in results:
        line = line.replace("<mark>", "")
        line = line.replace("</mark>", "")
        cleaned.append(line)

    return cleaned

def detect_log_format(line):
    line = line.strip()

    if not line:
        return "empty"
    if line.startswith("{") and line.endswith("}"):
        return "json"

    if " - - [" in line and '] \"' in line:
        return "apache"

    if any(level in line.upper() for level in [
        "CRITICAL",
        "ERROR",
        "WARNING",
        "INFO",
        "DEBUG",
        "TRACE",
    ]):
        return "standard"
    return "unknown"

def parse_json_log_line(line):
    try:
        data = json.loads(line)
    except json.JSONDecodeError:
        return line

    level = str(data.get("level", "")).upper()
    message = str(data.get("message", ""))

    if level and message:
        return f"{level} {message}"

    return ""

def parse_apache_log_line(line):
    parts = line.split('"')
    if len(parts) < 3:
        return line

    request_part = parts[1].strip()
    status_part = parts[2].strip().split()

    if not status_part:
        return line

    status_code = status_part[0]

    if status_code.startswith("5"):
        level = "ERROR"
    elif status_code.startswith("4"):
        level = "WARNING"
    else:
        level = "INFO"

    return f"{level} {request_part} {status_code}"
            
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        log_text = request.form.get("log_text", "")
        uploaded_file = request.files.get("log_file")
        source_name = "Pasted text"
        keyword = request.form.get("keyword", "")
        selected_levels = request.form.getlist("levels")
        case_sensitive = "case_sensitive" in request.form
        start_datetime_text = request.form.get("start_datetime", "")
        end_datetime_text = request.form.get("end_datetime", "")

        # 1. split lines
        if uploaded_file and uploaded_file.filename:
            source_name = uploaded_file.filename
            log_text = uploaded_file.read().decode("utf-8")

        if not log_text.strip():
            return render_template(
                "index.html",
                error_message="Please paste log text or upload a log file.",
            )

        lines = log_text.splitlines()
        total_input_lines = len(lines)

        detected_formats = [
            detect_log_format(line)
            for line in lines
            if line.strip()
        ]

        unique_formats = set(detected_formats)
        formats_found = sorted(unique_formats)

        format_counts = {
            format_name: detected_formats.count(format_name)
            for format_name in formats_found
        }

        dominant_format = (
            max(format_counts, key=format_counts.get)
            if format_counts
            else "unknown"
        )
        
        total_detected_formats = len(detected_formats)
        format_percentages = {
            format_name: (
                format_counts[format_name] / total_detected_formats * 100
                if total_detected_formats
                else 0
            )
            for format_name in formats_found
        }

        dominant_percentage = (
            format_percentages.get(dominant_format, 0)
            if format_percentages
            else 0
        )

        if len(unique_formats) > 1:
            detected_format = "mixed"
        elif detected_formats:
            detected_format = detected_formats[0]
        else:
            detected_format = "unknown"

        ignored_lines = 0
        blank_lines = 0
        invalid_json_lines = 0
        processed_lines = []
        
        for line in lines:
            if not line.strip():
                ignored_lines += 1
                blank_lines += 1
                continue

            line_format = detect_log_format(line)

            if line_format == "json":
                line = parse_json_log_line(line)

                if not line:
                    invalid_json_lines += 1
                    
            elif line_format == "apache":
                line = parse_apache_log_line(line)

            if line:
                processed_lines.append(line)
            else:
                ignored_lines += 1

        lines = processed_lines

        processed_line_count = len(processed_lines)

        processing_rate = (
            processed_line_count / total_input_lines * 100
            if total_input_lines
            else 0
        )
        
        start_datetime = None
        end_datetime = None

        if start_datetime_text:
            try:
                start_datetime = datetime.strptime(start_datetime_text, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return render_template(
                    "index.html",
                    error_message="Start datetime must be in format YYYY-MM-DD HH:MM:SS.",
                )

        if end_datetime_text:
            try:
                end_datetime = datetime.strptime(end_datetime_text, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return render_template(
                    "index.html",
                    error_message="End datetime must be in format YYYY-MM-DD HH:MM:SS.",
                )
            
        if start_datetime or end_datetime:
            filtered_lines = []

            for line in lines:
                line_datetime = parse_log_datetime(line)

                if line_datetime is None:
                    continue

                if start_datetime and line_datetime < start_datetime:
                    continue

                if end_datetime and line_datetime > end_datetime:
                    continue

                filtered_lines.append(line)

            lines = filtered_lines

        # 2. selected level filter
        if selected_levels:
            lines = [
                line for line in lines
                if any(level in line for level in selected_levels)
            ]

        # 3. keyword filter
        if keyword:
            if case_sensitive:                
                results = [
                    line for line in lines
                    if keyword in line
                ]
            else:
                results = [
                    line for line in lines
                    if keyword.lower() in line.lower()
                ]
        else:
            results = lines

        if keyword:
            results = [
                highlight_keyword(line, keyword, case_sensitive)
                for line in results
            ]
                              
        # 4. summary from final results
        summary = {
            "CRITICAL": 0,
            "ERROR": 0,
            "WARNING": 0,
            "INFO": 0,
            "DEBUG": 0,
            "TRACE": 0,
        }

        for line in results:
            for level in summary:
                if level in line:
                    summary[level] += 1

        global latest_results_text, latest_results_rows, history

        plain_results = [
            str(line).replace("<mark>", "").replace("</mark>", "")
            for line in results
        ]

        latest_results_text = "\n".join(plain_results)

        latest_results_rows =[]

        for line in plain_results:
            parts = line.split(" ", 3)

            if len(parts) == 4:
                timestamp = parts[0] + " " + parts[1]
                level = parts[2]
                message = parts[3]
            else:
                timestamp = ""
                level = ""
                message = line

            latest_results_rows.append([timestamp, level, message])
        
        history.append({
            "keyword": keyword.strip() or "Not set",
            "levels": ", ".join(selected_levels),       
            "matches": len(results),
            "searched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "results": results,
        })

        if len(history) > 100:
            history.pop(0)

        save_history(history)
            
        level_stats = {
            "CRITICAL": 0,
            "ERROR": 0,
            "WARNING": 0,
            "INFO": 0,
            "DEBUG": 0,
            "TRACE": 0,
        }

        for item in history:
            levels_text = item.get("levels", "")

            for level in level_stats:
                if level in levels_text:
                    level_stats[level] += 1

            total_searches = len(history)

            history_search = request.args.get("history_search", "")
            history_sort = request.args.get("history_sort", "newest")

            display_history = history

            if history_search:
                display_history = [
                    item for item in history
                    if (
                        history_search.lower() in item.get("keyword", "").lower()
                        
                        or history_search.lower() in item.get("levels", "").lower()
                    )                        
                ]

            if history_sort == "newest":
                display_history = list(reversed(display_history))

            elif history_sort == "oldest":
                display_history = display_history

            elif history_sort == "keyword":
                display_history = sorted(
                    display_history,
                    key=lambda item: item.get("keyword", "").lower()
                )
                
            successful_searches = sum(
                1 for item in history
                if item.get("matches", 0) > 0
            )

            total_matches_found = sum(
                item.get("matches", 0)
                for item in history
            )

            success_rate = 0

            if total_searches > 0:
                success_rate = round((successful_searches / total_searches) * 100)

            average_matches = 0

            if total_searches > 0:
                average_matches = round(total_matches_found / total_searches, 1)                    
                                    
            last_search_time = "N/A"

            if history:
                last_search_time = history[-1].get("searched_at") or "N/A"

            latest_keyword = "Not set"

            if history:
                latest_keyword = history[-1].get("keyword", "").strip()

                if not latest_keyword:
                    latest_keyword = "Not set"

            most_keyword = "N/A"

            if history:
                keyword_counts = {}

                for item in history:
                    key = item.get("keyword", "").strip()

                    if not key:
                        continue

                    keyword_counts[key] = keyword_counts.get(key, 0) + 1

                if keyword_counts:
                    most_keyword = max(
                        keyword_counts,
                        key=keyword_counts.get
                    )
                else:
                    most_keyword = "Not set"

        history_search = request.args.get("history_search", "")
        history_sort = request.args.get("history_sort", "newest")

        page = int(request.args.get("page", 1))
        per_page = 10

        display_history = history

        if history_search:
            display_history = [
                item for item in history
                if history_search.lower() in item.get("keyword", "").lower()
            ]

        if history_sort == "oldest":
            display_history = list(reversed(display_history))

        total_searches = len(history)

        chart_labels = []
        chart_values = []

        for item in display_history:
            label = item.get("keyword", "").strip()
            if not label:
                label = "Not set"

            chart_labels.append(label)
            chart_values.append(item.get("matches", 0))

        visible_level_counts = {
            "CRITICAL": 0,
            "ERROR": 0,
            "WARNING": 0,
            "INFO": 0,
            "DEBUG": 0,
            "TRACE": 0,
        }

        for item in display_history:
            item_levels = item.get("levels", "")

            for level in visible_level_counts:
                if level in item_levels:
                    visible_level_counts[level] +=1

                                        
        return render_template(
            "results.html",
            keyword=keyword,
            results=results,
            total=len(results),
            total_input_lines=total_input_lines,
            processed_line_count=processed_line_count,
            processing_rate=processing_rate,
            ignored_lines=ignored_lines,
            blank_lines=blank_lines,
            invalid_json_lines=invalid_json_lines,
            selected_levels=selected_levels,
            summary=summary,
            case_sensitive=case_sensitive,
            source_name=source_name,
            start_datetime_text=start_datetime_text,
            end_datetime_text=end_datetime_text,
            history=history,
            display_history=display_history[:10],
            page=1,
            total_pages=(len(display_history) +10 - 1) // 10,
            per_page=10,
            history_search=history_search,
            total_searches=total_searches,            
            successful_searches=successful_searches,
            total_matches_found=total_matches_found,
            average_matches=average_matches,
            success_rate=success_rate,
            most_keyword=most_keyword,
            latest_keyword=latest_keyword,
            last_search_time=last_search_time,
            level_stats=level_stats,
            history_sort=history_sort,
            visible_level_counts=visible_level_counts,
            chart_labels=chart_labels,
            chart_values=chart_values,
            detected_format=detected_format,
            formats_found=formats_found,
            format_counts=format_counts,
            dominant_format=dominant_format,
            dominant_percentage=dominant_percentage,
            format_percentages=format_percentages,            
        )

    return render_template("index.html")

@app.route("/view-history/<int:index>")
def view_history(index):
    if index < 0 or index >= len(history):
        return redirect("/filter-history")

    item = history[index]

    level_stats = {
        "CRITICAL": 0,
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "DEBUG": 0,
        "TRACE": 0,
    }

    for x in history:
        levels_text = x.get("levels", "")

        for level in level_stats:
            if level in levels_text:
                level_stats[level] += 1

    total_searches = len(history)
    successful_searches = sum(
        1 for x in history if x.get("matches", 0) > 0
    )

    total_matches_found = sum(
        x.get("matches", 0) for x in history
    )

    success_rate = 0
    if total_searches > 0:
        success_rate = round((successful_searches / total_searches) * 100)

    average_matches = 0
    if total_searches > 0:
        average_matches = round(total_matches_found / total_searches, 1)

    chart_labels = []
    chart_values = []

    for x in history:
        label = x.get("keyword", "").strip()

        if not label:
            label = "Not set"

        chart_labels.append(label)
        chart_values.append(x.get("matches", 0))

    most_keyword = "Not set"

    if history:
        keyword_counts = {}

        for x in history:
            key = x.get("keyword", "").strip()

            if not key:
                continue

            keyword_counts[key] = keyword_counts.get(key, 0) + 1

        if keyword_counts:
            most_keyword = max(
                keyword_counts,
                key=keyword_counts.get
            )

        display_history = history

    visible_level_counts = {
        "CRITICAL": 0,
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "DEBUG": 0,
        "TRACE": 0,
    }

    for item in display_history:
        item_levels = item.get("levels", "")

        for level in visible_level_counts:
            if level in item_levels:
                visible_level_counts[level] +=1
        
    return render_template(
        "results.html",
        keyword=item.get("keyword", ""),
        results=item.get("results", []),
        total=item.get("matches", 0),
        selected_levels=item.get("levels", "").split(", "),
        summary={},
        case_sensitive=False,
        source_name="History View",
        start_datetime_text="",
        end_datetime_text="",
        history=history,
        display_history=display_history,
        history_sort="newest",
        page=1,
        total_pages=1,
        per_page=10,
        history_search="",
        total_searches=total_searches,
        successful_searches=successful_searches,        
        total_matches_found=total_matches_found,
        success_rate=success_rate,
        average_matches=average_matches,
        latest_keyword=item.get("keyword", ""),
        last_search_time=item.get("searched_at", ""),
        level_stats=level_stats,
        most_keyword=most_keyword,
        visible_level_counts=visible_level_counts,
        
        chart_labels=chart_labels,
        chart_values=chart_values,
    )

@app.route("/delete-history/<int:index>", methods=["POST"])
def delete_history(index):
    if 0 <= index < len(history):
        history.pop(index)
        save_history(history)
        
    return redirect("/filter-history")

@app.route("/filter-history")
def filter_history():
    global latest_filtered_history
    last_search_time = "N/A"
    history_search = request.args.get("history_search", "")
    history_sort = request.args.get("history_sort", "newest")
    history_from = request.args.get("history_from", "")
    history_to = request.args.get("history_to", "")
    history_level =request.args.get("history_level", "")

    display_history = history

    if history_search:
        search = history_search.lower()
        display_history = [
            item
            for item in history
            if search in item.get("keyword", "").lower()                                                          
        ]

    if history_level:
        display_history = [
            item for item in display_history
            if history_level in item.get("levels", "")
        ]
    
    if history_from:
        display_history = [
            item for item in display_history
            if item.get("searched_at", "")[:10] >= history_from
        ]

    if history_to:
        display_history = [
            item for item in display_history
            if item.get("searched_at", "")[:10] <= history_to
        ]

    if history_sort == "newest":
        display_history = sorted(
            display_history,
            key=lambda item: item.get("searched_at", ""),
            reverse=True,
        )

    elif history_sort == "oldest":
        display_history = sorted(
            display_history,
            key=lambda item: item.get("searched_at", "")
        )

    elif history_sort == "keyword_asc":
        real_keywords = [
            item for item in display_history
            if item.get("keyword", "").strip()
            and item.get("keyword", "").strip().lower() !="not set"
        ]
    
        missing_keywords = [
            item for item in display_history
            if not item.get("keyword", "").strip()
            or item.get("keyword", "").strip().lower() == "not set"
        ]

        real_keywords = sorted(
            real_keywords,
            key=lambda item: item.get("keyword", "").strip().lower(),
        )
        display_history = real_keywords + missing_keywords

    elif history_sort == "keyword_desc":
        real_keywords = [
            item for item in display_history
            if item.get("keyword", "").strip()
            and item.get("keyword", "").strip().lower() != "not set"
        ]

        missing_keywords = [
            item for item in display_history
            if not item.get("keyword", "").strip()
            or item.get("keyword", "").strip().lower() == "not set"
        ]

        real_keywords = sorted(
            real_keywords,
            key=lambda item: item.get("keyword", "").strip().lower(),
            reverse=True
        )

        display_history = real_keywords + missing_keywords

    elif history_sort == "levels_high":
        severity_order = {
            "CRITICAL": 0,
            "ERROR": 1,
            "WARNING": 2,
            "INFO": 3,
            "DEBUG": 4,
            "TRACE": 5,
        }

        display_history = sorted(
            display_history,
            key=lambda item: min(
                (
                    severity_order[level]
                    for level in severity_order
                    if level in item.get("levels", "")
                ),
                default=99,
            ),
        )

    elif history_sort == "levels_low":
        severity_order = {
            "CRITICAL": 0,
            "ERROR": 1,
            "WARNING": 2,
            "INFO": 3,
            "DEBUG": 4,
            "TRACE": 5,
        }

        display_history = sorted(
            display_history,
            key=lambda item: min(
                (
                    severity_order[level]
                    for level in severity_order
                    if level in item.get("levels", "")
                ),
                default=99,
            ),
            reverse=True,
        )
                                
    elif history_sort == "matches_high": 
        display_history = sorted(
            display_history,
            key=lambda item: item.get("matches", 0),
            reverse=True
        )

    elif history_sort == "matches_low":
        display_history = sorted(
            display_history,
            key=lambda item: item.get("matches", 0)
        )

    stats_history = display_history

    visible_total_searches = len(stats_history)

    visible_unique_keywords = len(set(
        item.get("keyword", "").strip() or "Not set"
        for item in stats_history
    ))

    visible_total_matches = sum(
        item.get("matches", 0)
        for item in stats_history
    )

    visible_first_search = "N/A"
    visible_last_search = "N/A"

    search_times = [
        item.get("searched_at", "")
        for item in stats_history
        if item.get("searched_at", "")
    ]
    
    visible_first_search = min(search_times) if search_times else "N/A"
    visible_last_search = max(search_times) if search_times else "N/A"
                                             
    page = int(request.args.get("page", 1))
    per_page = 10

    total_pages = (len(display_history) + per_page - 1) // per_page

    start = (page - 1) * per_page
    end = start + per_page

    paged_history = display_history[start:end]
    latest_filtered_history = paged_history

    chart_counts = {}

    for item in display_history:
        label = item.get("keyword", "").strip()

        if not label:
            label = "Not set"

        chart_counts[label] = (
            chart_counts.get(label, 0)
            + item.get("matches", 0)
        )

    chart_labels = list(chart_counts.keys())
    chart_values = list(chart_counts.values())            
                                            
    level_stats = {
        "CRITICAL": 0,
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "DEBUG": 0,
        "TRACE": 0,
    }

    for item in stats_history:
        levels_text = item.get("levels", "")

        for level in level_stats:
            if level in levels_text:
                level_stats[level] += 1

    total_matches_found = sum(
        item.get("matches", 0)
        for item in stats_history
    )

    success_rate = 0
    if len(stats_history) > 0:
        success_rate = round((sum(1 for item in stats_history if item["matches"] > 0) / len(stats_history)) * 100)

    average_matches = 0
    if len(stats_history) > 0:
        average_matches = round(total_matches_found / len(stats_history), 1)

    latest_keyword = "Not set"
    if stats_history:
        latest_keyword = stats_history[0].get("keyword", "").strip()
        if not latest_keyword:
            latest_keyword = "Not set"

    last_searh_time = "N/A"
    if stats_history:
        last_search_time = stats_history[0].get("searched_at") or "N/A"

    most_keyword = "Not set"

    keyword_counts = {}

    for item in stats_history:
        key = item.get("keyword", "").strip()

        if not key:
            continue

        keyword_counts[key] = keyword_counts.get(key, 0) + 1

    if keyword_counts:
        most_keyword = max(
            keyword_counts,
            key=keyword_counts.get
        )

    visible_level_counts = {
        "CRITICAL": 0,
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "DEBUG": 0,
        "TRACE": 0,
    }

    for item in display_history:
        item_levels = item.get("levels", "")

        for level in visible_level_counts:
            if level in item_levels:
                visible_level_counts[level] +=1

    return render_template(
        "results.html",
        keyword="",
        results=[],
        total=0,
        selected_levels=[],
        summary={"ERROR": 0, "WARNING": 0, "INFO": 0, "DEBUG": 0},
        case_sensitive=False,
        source_name="History filter",
        start_datetime_text="",
        end_datetime_text="",
        history=history,
        display_history=paged_history,
        paged_history=paged_history,
        page=page,
        total_pages=total_pages,
        per_page=per_page,                        
        history_sort=history_sort,        
        history_search=history_search,
        history_from=history_from,
        history_to=history_to,
        history_level=history_level,
        visible_level_counts=visible_level_counts,
        total_searches=len(stats_history),
        successful_searches=sum(1 for item in stats_history if item["matches"] > 0),
        visible_total_searches=visible_total_searches,
        visible_unique_keywords=visible_unique_keywords,
        visible_total_matches=visible_total_matches,
        visible_first_search=visible_first_search,
        visible_last_search=visible_last_search,
        total_matches_found=total_matches_found,
        success_rate=success_rate,
        average_matches=average_matches,
        latest_keyword=latest_keyword,
        last_search_time=last_search_time,
        level_stats=level_stats,
        most_keyword=most_keyword,
        chart_labels=chart_labels,
        chart_values=chart_values,
    )

@app.route("/download")
def download_results():
    file_data = BytesIO()

    file_data.write(latest_results_text.encode("utf-8"))

    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_results.txt",
        mimetype="text/plain"
    )

@app.route("/download-backup")
def download_backup():
    backup_data = {
        "backup_version": 1,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "history": history,
    }

    file_data = BytesIO()
    file_data.write(
        json.dumps(backup_data, indent=2).encode("utf-8")
    )
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="web_log_analyzer_backup.json",
        mimetype="application/json"
    )

@app.route("/download-history-csv")
@app.route("/export-history-csv")
def download_history_csv():
    history_sort = request.args.get("history_sort", "newest")
    history_search = request.args.get(
        "history_search", ""
    ).strip().lower()        
    history_from = request.args.get("history_from", "").strip()
    history_to = request.args.get("history_to", "").strip()
    history_level = request.args.get("history_level", "").strip()
        
    export_history = list(history)

    if history_search:
        export_history = [
            item
            for item in export_history
            if history_search in item .get("keyword", "").lower()
        ]

    if history_from:
        export_history = [
            item
            for item in export_history
            if item .get("searched_at", "")[:10] >= history_from
        ]

    if history_to:
        export_history = [
            item
            for item in export_history
            if item .get("searched_at", "")[:10] <= history_to
        ]
    if history_level:
        export_history = [
            item
            for item in export_history
            if history_level in item.get("levels", "")
        ]
    severity_order = {
        "CRITICAL": 0,
        "ERROR": 1,
        "WARNING": 2,
        "INFO": 3,
        "DEBUG": 4,
        "TRACE": 5,
    }

    if history_sort == "oldest":
        export_history.sort(key=lambda item: item.get("searched_at", ""))

    elif history_sort == "newest":
        export_history.sort(
            key=lambda item: item.get("searced_at", ""),
            reverse=True,
        )

    elif history_sort == "matches_low":
        export_history.sort(
            key=lambda item: int(item.get("matches", 0) or 0)
        )

    elif history_sort == "matches_high":
        export_history.sort(
            key=lambda item: int(item.get("matches", 0) or 0),
            reverse=True,
        )

    elif history_sort == "keyword_asc":
        export_history.sort(
            key=lambda item: item.get("keyword", "").lower()
        )

    elif history_sort == "keyword_desc":
        export_history.sort(
            key=lambda item: item.get("keyword", "").lower(),
            reverse=True
        )

    elif history_sort in ("levels_high", "levels_low"):
        export_history.sort(
            key=lambda item: min(
                (
                    severity_order[level]
                    for level in severity_order
                    if level in item.get("levels", "")
                ),
                default=99,
            ),
            reverse=(history_sort == "levels_low"),
        )

    text_stream = StringIO()
    writer = csv.writer(text_stream)

    writer.writerow(["Keword", "Levels", "Matches", "Searched At"])
    for item in export_history:
        writer.writerow(
            [
                item.get("keyword", ""),
                item.get("levels", ""),
                item.get("matches", 0),
                item.get("searched_at", ""),
            ]
        )

    file_data = BytesIO(text_stream.getvalue().encode("utf-8-sig"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="filtered_history.csv",
        mimetype="text/csv",
    )                        

@app.route("/download-csv")
def download_csv():
    file_data = BytesIO()

    text_stream = "\n".join(
        [",".join(["Timestamp", "Level", "Message"])] +
        [",".join(row) for row in latest_results_rows]
    )

    file_data.write(text_stream.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_results.csv",
        mimetype="text/csv"
    )

@app.route("/clear-history", methods=["POST"])
def clear_hisrory():
    global history

    history.clear()
    save_history(history)

    return redirect("/")

@app.route("/download-history")
def download_history():
    return send_file(
        "history.json",
        as_attachment=True,
        download_name="history.json"
    )

@app.route("/download-filtered-history")
def download_filtered_history():
    file_data = BytesIO()

    export_data = {
        "total_searches": len(latest_filtered_history),
        "successful_searches": sum(
            1 for item in latest_filtered_history
            if item["matches"] > 0
        ),
        "most_searched_keyword": (
            latest_filtered_history[0]["keyword"]
            if latest_filtered_history
            else "N/A"
        ),
        "history": latest_filtered_history
    }

    text_stream = json.dumps(
        export_data,
        indent=2
        
    )

    file_data.write(text_stream.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="filtered_history.json",
        mimetype="application/json"
    )

@app.route("/download-stats")
def download_stats():
    file_data = BytesIO()

    stats_data = {
        "total_searches": len(history),
        "successful_searches": sum(
            1 for item in history
            if item["matches"] > 0
        ),
        "total_matches_found": sum(
            item["matches"] for item in history
        ),
        "most_searched_keyword": "N/A",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    if history:
        keyword_counts = {}

        for item in history:
            key = item["keyword"]
            keyword_counts[key] = keyword_counts.get(key, 0) + 1

        stats_data["most_searched_keyword"] = max(
            keyword_counts,
            key=keyword_counts.get
        )

    text_stream = json.dumps(stats_data, indent=2)

    file_data.write(text_stream.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="history_stats.json",
        mimetype="application/json"
    )
@app.route("/download-stats-csv")
def download_stats_csv():

    csv_data = StringIO()

    csv_data.write("Metric,Value\n")
    csv_data.write(f"Total searches,{len(history)}\n")

    successful_searches = sum(
        1 for item in history
        if item.get("matches", 0) > 0
    )

    csv_data.write(f"Successful searches,{successful_searches}\n")

    total_matches = sum(
        item.get("matches", 0)
        for item in history
    )

    csv_data.write(f"Total matches found,{total_matches}\n")

    keyword_counts = {}

    for item in history:
        keyword = item . get("keyword", "").strip()
        if keyword:
            keyword_counts[keyword] = keyword_counts.get(keyword, 0) + 1

    most_searched_keyword = (
        max(keyword_counts, key=keyword_counts.get)
        if keyword_counts
        else "Not set"
    )

    csv_data.write(f"Most searched keyword, {most_searched_keyword}\n")
    csv_data.write(f"Generated at, {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    file_data = BytesIO()
    file_data.write(csv_data.getvalue().encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="history_stats.csv",
        mimetype="text/csv"
    )

@app.route("/download-stats-excel")
def download_stats_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stats"
    sheet.freeze_panes = "A2"

    sheet.append(["Metric", "Value"])

    successful_searches = sum(
        1
        for item in history
        if item .get("matches", 0) > 0
    )

    total_matches = sum(
        item.get("matches", 0)
        for item in history
    )

    keyword_counts = {}

    for item in history:
        keyword = item.get("keyword", "").strip()
        if keyword:
            keyword_counts[keyword] = keyword_counts.get(keyword, 0) + 1

    most_searched_keyword = (
        max(keyword_counts, key=keyword_counts.get)
        if keyword_counts
        else "Not set"
    )

    sheet.append(["Total searches", len(history)])
    sheet.append(["Successful searches", successful_searches])
    sheet.append(["Total matches found", total_matches])
    sheet.append(["Most searched keyword", most_searched_keyword])
    sheet.append(["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 2

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )

    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="history_stats.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/import-history", methods=["GET", "POST"])
def import_hiatory():
    if request.method == "POST":
        uploaded_file = request.files.get("history_file")

        if not uploaded_file or not uploaded_file.filename.lower().endswith(".json"):
            return "Please select a valid JSON file."

        try:
            imported_history = json.load(uploaded_file)
        except (json.JSONDecodeError, UnicodeDecodeError):
            return "The selected file does not contain valid JSON."

        if isinstance(imported_history, dict):
            if "history" not in imported_history:
                return "Invalid backup file: missing history data."
            imported_history = imported_history["history"]

        if not isinstance(imported_history, list):
            return "Invalid history file: the JSON must contain a list."

        required_fields = {"keyword", "levels", "matches", "searched_at", "results"}

        for item in imported_history:
            if not isinstance(item, dict):
                return "Invalid history file: every record must be an object."

            if not required_fields.issubset(item):
                return "Invaltd history file: one or more records are missing required fields."

        new_records = []
        duplicate_count = 0

        for imported_item in imported_history:
            if imported_item in history:
                duplicate_count += 1
            else:
                new_records.append(imported_item)        
                    
        history.extend(new_records)

        return (
            f"Imported {len(new_records)} new history records.<br>"
            f"Skipped {duplicate_count} duplicate records.<br><br>"
            f'<a href="/filter-history"> Back to History</a>'
        )
                                   
    return render_template("import_history.html")
                                                                    
@app.route("/download-history-excel")
def download_history_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "History"
    sheet.freeze_panes = "A2"

    sheet.append(["Keyword", "Levels", "Matches", "Searched At", "Results"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for cell in sheet[1]:
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )

    for item in history:
        sheet.append([
            item.get("keyword", ""),
            item.get("levels", ""),
            item.get("matches", ""),
            item.get("searched_at", ""),
            "\n".join(clean_export_results(item.get("results", [])))
        ])
    
        sheet.column_dimensions["E"].width = 60

        for cell in sheet["E"]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in sheet.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(
                        fill_type="solid",
                        start_color="EDEDED",
                        end_color="EDEDED"
                    )
                    
        sheet.auto_filter.ref = sheet.dimensions
                
    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.freeze_panes = "A2"
    
    summary_sheet.append(["Summary Metric", "value"])

    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )
            
    stats_sheet = workbook.create_sheet("Stats")
    stats_sheet.freeze_panes = "A2"

    stats_sheet.append(["Metric", "Value"])
    for cell in stats_sheet[1]:
        cell.font = Font(bold=True)

    for cell in stats_sheet[1]:
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3",
        )
    
    stats_sheet.append(["Total searches", len(history)])

    stats_sheet.append([
        "Successful searches",
        sum(1 for item in history if item.get("matches", 0) > 0)
    ])
        
    most_keyword = "N/A"

    if history:
        keyword_counts = {}

        for item in history:
            key = item.get("keyword", "").strip()
            if not key:
                key = "Not set"

            keyword_counts[key] = keyword_counts.get(key, 0) + 1

        most_keyword = max(keyword_counts, key=keyword_counts.get)

        summary_sheet.append(["Total searches", len(history)])

        summary_sheet.append([
            "Successful searches",
            sum(1 for item in history if item.get("matches", 0) > 0)
        ])

        summary_sheet.append([
            "Total matches found",
            sum(item.get("matches", 0) for item in history)
        ])

        summary_sheet.append([
            "Most searched keyword",
            most_keyword
        ])

    stats_sheet.append(["Most searched keyword", most_keyword])

    stats_sheet.append([])

    stats_sheet.append(["Level", "Count"])
    for cell in stats_sheet[6]:
        cell.font = Font(bold=True)

    for cell in stats_sheet[6]:
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3",
        )
    
    for level in ["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG", "TRACE"]:
        count = 0

        for item in history:
            if level in item.get("levels", ""):
                count += 1

        stats_sheet.append([level, count])

    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                cell_value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(cell_value))

            worksheet.column_dimensions[column_letter].width = max_length + 2

    stripe_fill = PatternFill(
        fill_type="solid",
        start_color="F2F2F2",
        end_color="F2F2F2",
    )
    critical_fill = PatternFill(
        start_color="FF9999",
        end_color="FF9999",
        fill_type="solid",
    )

    error_fill = PatternFill(
        start_color="FFCC99",
        end_color="FFCC99",
        fill_type="solid",
    )

    warning_fill = PatternFill(
        start_color="FFFF99",
        end_color="FFFF99",
        fill_type="solid",
    )

    info_fill = PatternFill(
        start_color="CCFFFF",
        end_color="CCFFFF",
        fill_type="solid",
    )

    debug_fill = PatternFill(
        start_color="DDDDDD",
        end_color="DDDDDD",
        fill_type="solid",
    )

    trcae_fill = PatternFill(
        start_color="CCFFCC",
        end_color="CCFFCC",
        fill_type="solid",
    )

    for row in sheet.iter_rows(min_row=2):
        if row[0].row % 2 ==0:
            for cell in row:
                cell.fill = stripe_fill

    for row in sheet.iter_rows(min_row=2):
        levels_cell = row[1]
        levels_text = str(levels_cell.value or "")

        if "CRITICAL" in levels_text:
            levels_cell.fill = critical_fill
        elif "ERROR" in levels_text:
            levels_cell.fill = error_fill
        elif "WARNING" in levels_text:
            levels_cell.fill = warning_fill
        elif "INFO" in levels_text:
            levels_cell.fill = info_fill
        elif "DEBUG" in levels_text:
            levels_cell.fill = debug_fill
        elif "TRACE" in levels_text:
            levels_cell.fill = trace_fill

    summary_sheet.auto_filter.ref = summary_sheet.dimensions            
    stats_sheet.auto_filter.ref = stats_sheet.dimensions

    for column in stats_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        stats_sheet.column_dimensions[column_letter].width = max_length + 2

    
    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="history.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/export-history-xlsx")
@app.route("/download-filtered-history-excel")
def download_filtered_history_excel():
    history_search = request.args.get("history_search", "")
    history_sort = request.args.get("history_sort", "newest")
    history_from = request.args.get("history_from", "")
    history_to = request.args.get("history_to", "")
    history_level = request.args.get("history_level", "")

    display_history = list(history)

    if history_search:
        display_history = [
            item
            for item in disply_history
            if (
                history_search.lower() in item.get("keyword", "").lower()
                or history_search.lower() in item.get("levels", "").lower()
            )
        ]

    if history_from:
        display_history = [
            item for item in display_history
            if item.get("searched_at", "")[:10] >= history_from
        ]

    if history_to:
        display_history = [
            item for item in display_history
            if item.get("searched_at", "")[:10] <= history_to
        ]

    if history_level:
        display_history = [
            item
            for item in display_history
            if history_level in item.get("levels", "")
        ]

    severity_order = { 
        "CRITICAL": 0,
        "ERROR": 1,
        "WARNING": 2,
        "INFO": 3,
        "DEBUG": 4,
        "TRACE": 5,
    }
                                            
    if history_sort == "newest":
        display_history.sort(
            key=lambda item: item.get("searched_at", ""),
            reverse=True,
        )

    elif history_sort == "oldest":
        display_history.sort(
            key=lambda item: item.get("searched_at", "")
        )

    elif history_sort == "matches_high":
        display_history.sort(
            key=lambda item: int(item.get("matches", 0) or 0),
            reverse=True,
        )

    elif history_sort == "matches_low":
        display_history.sort(
            key=lambda item: int(item.get("matches", 0) or 0)
        )

    elif history_sort == "keyword_asc":
        display_history.sort(
            key=lambda item: item.get("keyword", "").lower()
        )

    elif history_sort == "keyword_desc":
        display_history.sort(
            key=lambda item: item.get("keyword", "").lower(),
            reverse=True,
        )

    elif history_sort in ("levels_high", "levels_low"):
        display_history.sort(
            key=lambda item: min(
                (
                    severity_order[level]
                    for level in severity_order
                    if level in item.get("levels", "")
                ),
                default=99,
            ),
            reverse=(history_sort == "levels_low"),
        )
                            
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Filtered History"

    sheet.append(["Keyword", "Levels", "Matches", "Searched At"])

    for item in display_history:
        sheet.append([
            item.get("keyword", ""),
            item.get("levels", ""),
            item.get("matches", ""),
            item.get("searched_at", "")
        ])

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            cell_value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(cell_value))

    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 45
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 25

    for cell in sheet[1]:
        cell.font =Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3",
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.freeze_panes = "A2"

    summary_sheet.append(["Summary Metric", "Value"])

    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3",
        )

    summary_sheet.append(["Total searches", len(display_history)])

    summary_sheet.append([
        "Successful searches",
        sum(
            1
            for item in display_history
            if item.get("matches", 0) > 0
        ),
    ])

    summary_sheet.append([
        "Total matches found",
        sum(
            int(item.get("matches", 0) or 0)
            for item in display_history
        ),
    ])

    total_matches = sum(
        int(item.get("matches", 0) or 0)
        for item in display_history
    )

    average_matches = (
        round(total_matches / len(display_history), 2)
        if display_history else 0
    )
    
    summary_sheet.append([
        "Average matches per search",
        average_matches
    ])

    highest_matches = max(
        (
            int(item.get("matches", 0) or 0)
            for item in display_history
        ),
        default=0,
    )

    summary_sheet.append([
        "Highest matches in one search",
        highest_matches,
    ])

    earliest_search = min(
        (
            item.get("searched_at", "")
            for item in display_history
            if item.get("searched_at")
        ),
        default="N/A",
    )

    summary_sheet.append([
        "Earliest search",
        earliest_search,
    ])

    latest_search = max(
        (
            item.get("searched_at", "")
            for item in display_history
            if item .get("searched_at")
        ),
        default="N/A",
    )
    summary_sheet.append([
        "Latest search",
        latest_search,
    ])

    most_keyword = "N/A"

    if display_history:
        keyword_counts = {}

        for item in display_history:
            key = item.get("keyword", "").strip()
            if not key:
                key = "Not set"
            keyword_counts[key] = keyword_counts.get(key, 0) + 1

        most_keyword = max(
            keyword_counts,
            key=keyword_counts.get,
            default="N/A"
        )

    summary_sheet.append([
        "Most searched keyword",
        most_keyword
    ])

    most_common_level = "N/A"

    if display_history:
        level_counts = {
            "CRITICAL": 0,
            "ERROR": 0,
            "WARNING": 0,
            "INFO": 0,
            "DEBUG": 0,
            "TRACE": 0,
        }

        for item in display_history:
            level_text = item.get("levels", "")

            for level in level_counts:
                if level in level_text:
                    level_counts[level] += 1
        most_common_level = max(
            level_counts,
            key=level_counts.get,
            default = "N/A",
        )
    summary_sheet.append([
        "Most common log level",
        most_common_level,
    ])

    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 27

    for cell in summary_sheet["B"]:
        cell.alignment = Alignment(horizontal="center")

    charts_sheet = workbook.create_sheet("Charts")
    charts_sheet.append(["Log_level", "Count"])

    for cell in charts_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3",
        )

    for level in [
        "CRITICAL",
        "ERROR",
        "WARNING",
        "INFO",
        "DEBUG",
        "TRACE",
    ]:
        if history_level:
            count = len(display_history) if level == history_level else 0
        else:
            count = level_counts.get(level, 0)

        charts_sheet.append([
            level,
            count,
        ])

    charts_sheet.column_dimensions["A"].width = 15
    charts_sheet.column_dimensions["B"].width = 12

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Log Level Counts"            
    bar_chart.legend = None
    bar_chart.varyColors = False

    data = Reference(
        worksheet=charts_sheet,
        min_col=2,
        min_row=1,
        max_row=7,
    )

    categories = Reference(
        worksheet=charts_sheet,
        min_col=1,
        min_row=2,
        max_row=7,
    )

    bar_chart.add_data(
        data,
        titles_from_data=True,
    )

    bar_chart.set_categories(categories)

    bar_chart.x_axis.delete = False
    bar_chart.x_axis.tickLblPos = "nextTo" 
    

    bar_chart.dLbls = DataLabelList()
    bar_chart.dLbls.showVal = True
    bar_chart.dLbls.showCatName = False
    bar_chart.dLbls.showSerName = False
    bar_chart.dLbls.showLegendKey = False
    bar_chart.dLbls.dLblPos = "outEnd"

    highest_level_count = max(
        level_counts.values(),
        default=0,
    )

    bar_chart.y_axis.scaling.min =0
    bar_chart.y_axis.scaling.max = highest_level_count + 5
    bar_chart.y_axis.majorUnit = 5
    
    bar_chart.height = 8
    bar_chart.width = 14

    charts_sheet.add_chart(bar_chart, "D2")
            
    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0) 

    return send_file(
        file_data,
        as_attachment=True,
        download_name="filtered_history.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/export-history-pdf")
def export_history_pdf():
    file_data = BytesIO()

    history_search = request.args.get("history_search", "")
    history_sort = request.args.get("history_sort", "newest")
    history_from = request.args.get("history_from", "")
    history_to = request.args.get("history_to", "")
    history_level = request.args.get("history_level", "")

    display_history = list(history)

    if history_search:
        display_history = [
            item
            for item in display_history
            if history_search.lower() in item.get("keyword", "").lower()
        ]

    if history_level:
        display_history = [
            item
            for item in display_history
            if history_level.upper()
            in item.get("levels", "").upper()
        ]


    if history_from:
        display_history = [
            item
            for item in display_history
            if item.get("searched_at", "") >= history_from
        ]

    if history_to:
        display_history = [
            item
            for item in display_history
            if item.get("searched_at", "") <= history_to + " 23:59:59"
        ]
        
    doc = SimpleDocTemplate(
        file_data,
        pagesize=letter,
    )

    styles = getSampleStyleSheet()

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total_visible_searches = len(display_history)
    total_visible_matches = sum(
        item["matches"] for item in display_history
    )

    successful_searches = sum(
        1 for item in display_history if item.get("matches", 0) > 0
    )

    success_rate = (
        successful_searches / total_visible_searches * 100
        if total_visible_searches
        else 0
    )

    average_matches = (
        total_visible_matches / total_visible_searches
        if total_visible_searches
        else 0
    )

    search_times = [
        item.get("searched_at", "")
        for item in display_history
        if item.get("searched_at", "")
    ]

    first_visible_search = min(search_times) if search_times else "N/A"
    last_visible_search = max(search_times) if search_times else "N/A"

    visible_levels = sorted(
        {
            level.strip()
            for item in display_history
            for level in item.get("levels", "").split(",")
            if level.strip()
        }
    )

    keyword_counts = {}

    for item in display_history:
        keyword = item.get("keyword", "").strip()

        if keyword and keyword.lower() !="not set":
            keyword_counts[keyword] = keyword_counts.get(keyword, 0) + 1

    top_keyword = (
        max(keyword_counts, key=keyword_counts.get)
        if keyword_counts
        else "N/A"                                        
    )

    larest_item = max(
        display_history,
        key=lambda item: item.get("searched_at", "")
    ) if display_history else None 

    latest_keyword = (
        display_history[-1].get("keyword", "N/A")
        if display_history
        else "N/A"
    )

    if not latest_keyword:
        latest_keyword = "Not set"

    chart_counts = {}
    for item in display_history:
        keyword = item.get("keyword", "").strip()

        if not keyword:
            keyword = "Not set"

        chart_counts[keyword] = (
            chart_counts.get(keyword, 0)
            + item.get("matches", 0)
        )

    chart_labels = list(chart_counts.keys())
    chart_values = list(chart_counts.values())

    level_counts = {
        "CRITICAL": 0,
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "DEBUG": 0,
        "TRACE": 0,
    }

    for item in display_history:
        item_levels = item.get("levels", "")

        for level in level_counts:
            if history_level:
                if level == history_level:
                    level_counts[level] += 1
            elif level in item_levels:
                level_counts[level] += 1

    level_chart_labels = list(level_counts.keys())
    level_chart_values = list(level_counts.values())

    date_counts = {}

    for item in display_history:
        searched_at = item.get("searched_at", "")
        search_date = searched_at[:10]

        if search_date:
            date_counts[search_date] = date_counts.get(search_date, 0) + 1

    date_chart_labels = [
        datetime.strptime(date, "%Y-%m-%d").strftime("%b %d")
        for date in date_counts.keys()
    ]
    date_chart_values = list(date_counts.values())

    matches_by_date = {}

    for item in display_history:
        searched_at = item.get("searched_at", "")
        search_date = searched_at[:10]

        if search_date:
            matches_by_date[search_date] = (
                matches_by_date.get(search_date, 0)
                + item.get("matches", 0)
            )

    matches_date_labels = [
        datetime.strptime(date, "%Y-%m-%d").strftime("%b %d")
        for date in matches_by_date.keys()
    ]

    matches_date_values = list(matches_by_date.values())

    chart_drawing = Drawing(500, 220)

    bar_chart = VerticalBarChart()
    bar_chart.x = 50
    bar_chart.y = 40
    bar_chart.height = 140
    bar_chart.width = 400

    bar_chart.bars[0].fillColor = colors.darkblue

    bar_chart.data = [chart_values]
    bar_chart.categoryAxis.categoryNames = chart_labels

    bar_chart.valueAxis.valueMin = 0
    bar_chart.barLabels.nudge = 7
    bar_chart.barLabels.fontSize = 8
    bar_chart.barLabelFormat = "%d"

    chart_drawing.add(bar_chart)

    level_chart_drawing = Drawing(500, 220)

    level_bar_chart = VerticalBarChart()
    level_bar_chart.x = 50
    level_bar_chart.y = 40
    level_bar_chart.height = 140
    level_bar_chart.width = 400

    level_bar_chart.bars[0].fillColor = colors.darkblue

    level_bar_chart.data = [level_chart_values]
    level_bar_chart.categoryAxis.categoryNames = level_chart_labels

    level_bar_chart.valueAxis.valueMin = 0

    level_bar_chart.barLabels.nudge = 7
    level_bar_chart.barLabels.fontSize = 8
    level_bar_chart.barLabelFormat = "%d"

    level_chart_drawing.add(level_bar_chart)

    date_chart_drawing = Drawing(500, 220)

    date_bar_chart = VerticalBarChart()
    date_bar_chart.x = 50
    date_bar_chart.y = 40
    date_bar_chart.height = 140
    date_bar_chart.width = 400

    date_bar_chart.bars[0].fillColor = colors.darkblue

    date_bar_chart.data = [date_chart_values]
    date_bar_chart.categoryAxis.categoryNames = date_chart_labels

    date_bar_chart.valueAxis.valueMin = 0

    date_bar_chart.barLabels.nudge = 7
    date_bar_chart.barLabels.fontSize = 8
    date_bar_chart.barLabelFormat = "%d"

    date_chart_drawing.add(date_bar_chart)

    matches_date_chart_drawing = Drawing(500, 220)

    matches_date_bar_chart = VerticalBarChart()
    matches_date_bar_chart.x = 50
    matches_date_bar_chart.y = 40
    matches_date_bar_chart.height = 140
    matches_date_bar_chart.width = 400

    matches_date_bar_chart.bars[0].fillColor = colors.darkblue

    matches_date_bar_chart.data = [matches_date_values]
    matches_date_bar_chart.categoryAxis.categoryNames = matches_date_labels

    matches_date_bar_chart.valueAxis.valueMin = 0

    matches_date_bar_chart.barLabels.nudge = 7
    matches_date_bar_chart.barLabels.fontSize = 8
    matches_date_bar_chart.barLabelFormat = "%d"

    matches_date_chart_drawing.add(matches_date_bar_chart)
        
    summary_data = [
        ["Metric", "Value"],
        ["Total Visible Searches", str(total_visible_searches)],
        ["Successful Searches", str(successful_searches)],
        ["Success Rate", f"{success_rate:.0f}%"],
        ["Top Keyword", top_keyword],
        ["Latest Keyword", latest_keyword],
        ["Unique Keywords", str(len(set(item["keyword"] for item in display_history)))],
        ["Total Visible Matches", str(sum(item["matches"] for item in display_history))],
        ["Average Matches per Search", f"{average_matches:.1f}"],
        ["First Visible Search", first_visible_search],
        ["Last Visible Search", last_visible_search],
        ["Visible History Levels", ", ".join(visible_levels)],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[200, 320],
    )

    summary_table.halign = "CENTER"

    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    filters_data = [
        ["Filter", "Value"],
        ["Keyword", history_search or "All"],
        ["Level", history_level or "All Levels"],
        ["From Date", history_from or "Not set"],
        ["To Date", history_to or "Not set"],
        ["Sort Order", history_sort.replace("_", " ").title()],
    ]

    filters_table = Table(
        filters_data,
        colWidths=[150, 370],
    )

    filters_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements = [
        Paragraph("Web Log Analyzer PDF Export", styles["Title"]),                                
        Paragraph(f"Generated: {generated_at}", styles["Normal"]),
        Paragraph("Filters Applied", styles["Heading2"]),
        filters_table,
        Spacer(1, 12),
        Paragraph("Summary", styles["Heading2"]),
        summary_table,
        Spacer(1, 18),
        KeepTogether([
            Paragraph("Keyword Matches Chart", styles["Heading2"]),
            chart_drawing,
        ]),

        Spacer(1, 18),
        KeepTogether([
            Paragraph("Log Level Counts Chart", styles["Heading2"]),
            level_chart_drawing,
        ]),

        Spacer(1, 18),
        KeepTogether([
            Paragraph("Searches by Date Chart", styles["Heading2"]),
            date_chart_drawing,
        ]),

        Spacer(1, 18),
        KeepTogether([
            Paragraph("Matches by Date Chart", styles["Heading2"]),
            matches_date_chart_drawing,
        ]),
    ]

    doc.build(elements)

    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="filtered_history.pdf",
        mimetype="application/pdf",
    )
    
@app.route("/download-filtered-history-csv")
def download_filtered_history_csv():
    history_search = request.args.get("history_search", "")
    history_sort = request.args.get("history_sort", "newest")
    history_from = request.args.get("history_from", "")
    history_to = request.args.get("history_to", "")
        
    display_history = history

    if history_search:
        search = history_search.lower()
        display_history = [
            item for item in history
            if (
                search in item.get("keyword", "").lower()
                or search in item.get("searched_at", "").lower()
            )
        ]

    if history_from:
        display_history= [
            item for item in display_history
            if item.get("searched_at", "")[:10] >= history_from
        ]

    if history_to:
        display_history = [
            item for item in display_history
            if item .get("searched_at", "")[:10] <= history_to
        ]
        
    if history_sort == "newest":
        display_history = list(reversed(display_history))

    elif history_sort == "oldest":
        display_history = display_history

    elif history_sort == "keyword":
        display_history = sorted(
            display_history,
            key=lambda item: item.get("keyword", "").lower()
        )

    file_data = BytesIO()

    rows = ["Keyword,Levels,Matches,Searched At"]

    for item in display_history:
        rows.append(
            f'"{item.get("keyword", "")}","{item.get("levels", "")}","{item.get("matches", "")}","{item.get("searched_at", "")}"'
        )

    text_stream = "\n".join(rows)

    file_data.write(text_stream.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="filtered_history.csv",
        mimetype="text/csv"
    )
    
@app.route("/download-filtered-history-json")
def download_filtered_history_json():
    history_search = request.args.get("history_search", "")
    history_sort = request.args.get("history_sort", "newest")
    history_from = request.args.get("history_from", "")
    history_to = request.args.get("history_to", "")

    display_history = history

    if history_search:
        display_history = [
            item for item in history
            if (
                history_search.lower() in item.get("keyword", "").lower()
                or history_search.lower() in item.get("levels", "").lower()
            )
        ]

    if history_from:
        display_history = [
            item for item in display_history
            if item.get("searched_at", "")[:10] >= history_from
        ]

    if history_to:
        display_history = [
            item for item in display_history
            if item.get("searched_at", "")[:10] <= history_to
        ]

    if history_sort == "newest":
        display_history = list(reversed(display_history))

    elif history_sort == "oldest":
        dispaly_history = display_history

    elif history_sort == "keyword":
        display_history = sorted(
            display_history,
            key=lambda item: item.get("keword", "").lower()
        )
    file_data = BytesIO()

    export_history = []

    for item in display_history:
        new_item = item.copy()

        if not new_item.get("keyword", "").strip():
            new_item["keyword"] = "Not set"

        new_item["results"] = clean_export_results(
            new_item.get("results", [])
        )

        export_history.append(new_item)
        
    text_stream = json.dumps(export_history, indent=2)

    file_data.write(text_stream.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="filtered_history.json",
        mimetype="application/json"
    )

@app.route("/download-stats-json")
def download_stats_json():
    print(history)

    top_kayword = "Not set"
    keyword_counts = {}

    for item in history:
        key = item.get("keyword", "").strip()

        if not key:
            key = "Not set"

        keyword_counts[key] = keyword_counts.get(key, 0) + 1

    if keyword_counts:
        top_keyword = max(
            keyword_counts,
            key=keyword_counts.get
        )

    latest_keyword = "Not set"
    last_search = "Not set"

    if history:
        latest = history[-1]
        latest_keyword = latest.get("keyword", "").strip()

        if not latest_keyword:
            latest_keyword = "Not set"

        last_search =latest.get("searched_at", "Not set")

    level_stats = {
        "CRITICAL": 0,
        "ERROR": 0,
        "WARNING": 0,
        "INFO": 0,
        "DEBUG": 0,
        "TRACE": 0,
    }
    for item in history:
        levels_text = item.get("levels", "")

        for level in level_stats:
            if level in levels_text:
                level_stats[level] +=1
        
    stats = {
        "total_searches": len(history),
        "successful_searches": sum(
            1 for item in history if item.get("matches", 0) > 0
        ),
        "success_rate": round(
            (sum(1 for item in history if item.get("matches", 0) > 0) / len(history)) * 100
        ) if history else 0,
        "total_matches": sum(
            item.get("matches", 0) for item in history
        ),
        "average_matches": round(
            sum(item.get("matches", 0) for item in history) / len(history),
            1
        ) if history else 0,
        "top_keyword": top_keyword,
        "latest_keyword": latest_keyword,
        "last_search": last_search,
        "level_stats": level_stats
    }

    file_data = BytesIO()
    text_stream = json.dumps(stats, indent=2)

    file_data.write(text_stream.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="history_stats.json",
        mimetype="application/json"
    )

@app.route("/download-analysis-txt")
def download_analysis_txt():
    lines = []
    if history:
        latest = history[-1]

        lines.append(f"Keyword: {latest.get('keyword', 'Not set')}")
        lines.append(f"Levels: {latest.get('levels', 'All')}")
        lines.append(f"Matches: {latest.get('matches', 0)}")
        lines.append("")
        lines.append("-" * 40)

        for line in latest.get("results", []):
            line = line.replace("<msrks>", "")
            line = line.replace("</marks>", "")
            lines.append(line)

    text = "\n".join(lines)
                
    file_data = BytesIO()
    file_data.write(text.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_results.txt",
        mimetype="text/plain"
    )

@app.route("/download-analysis-csv")
def download_analysis_csv():
    file_data = BytesIO()
    text_stream = StringIO()
    writer = csv.writer(text_stream)

    writer.writerow(["Level", "Message"])

    if history:
        latest = history[-1]

        for line in latest.get("results", []):
            line = line.replace("<marks>", "")
            line = line.replace("</marks>", "")

            parts = line.split(" ", 1)

            if len(parts) == 2:
                level = parts[0]
                message = parts[1]
            else:
                level = "UNKNOWN"
                message = line

            writer.writerow([level, message])
    file_data.write(text_stream.getvalue().encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_results.csv",
        mimetype="text/csv"
    )

@app.route("/download-analysis-json")
def download_analysis_json():
    latest = history[-1] if history else {}

    results = []

    for line in latest.get("results", []):
        line = line.replace("<mark>", "")
        line = line.replace("</mark>", "")

        parts = line.split(" ", 1)

        if len(parts) == 2:
            level = parts[0]
            message = parts[1]
        else:
            level = "UNKNOWN"
            message = line

        results.append({
            "level": level,
            "message": message
        })
        print("JSON results count:", len(results))

    data = {
        "keyword": latest.get("keyword", "Not set"),
        "levels": latest.get("levels", ""),
        "matches": latest.get("matches", 0),
        "results": results
    }
        
             
    file_data = BytesIO()
    text_stream = json.dumps(data, indent=2)

    file_data.write(text_stream.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_results.json",
        mimetype="application/json"
    )

@app.route("/download-analysis-html")
def download_analysis_html():
    latest = history[-1] if history else {}

    keyword = latest.get("keyword", "Not set")
    levels = latest.get("levels", "")
    matches = latest.get("matches", 0)

    result_lines = ""
    for line in latest.get("results", []):
        line = line.replace("<marks>", "")
        line = line.replace("</marks>", "")
        if line.startswith("CRITICAL"):
            css_class = "critical-report"
        elif line.startswith("ERROR"):
            css_class = "error-report"
        elif line.startswith("WARNING"):
            css_class = "warning-report"
        elif line.startswith("INFO"):
            css_class = "info-report"
        elif line.startswith("DEBUG"):
            css_class = "debug-report"
        elif line.startswith("TRACE"):
            css_class = "trace-report"
        else:
            css_class = ""
            
        result_lines += f'<li><pre class="{css_class}">{line}</pre></li>'
        
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Analysis Report</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                background-color: #f4f6f8;
                padding: 30px;
            }}
            .report {{
                max-width: 900px;
                margin: auto;
                background: white;
                border-radius: 10px;
            }}

            pre {{
                padding: 10px;
                border-left: 5px solid #555;
                background-color: #f1f1f1;
                border-radius: 6px;
            }}

            .critical-report {{
                border-left-color: #b00020;
                background-color: #ffe5e5;
            }}

            .error-report {{
                border-left-color: #d32f2f;
                background-color: #fff0f0;
            }}

            .warning-report {{
                border-left-color: #f9a825;
                background-color: #fff8d6;
            }}

            .info-report {{
                border-left-color: #1976d2;
                background-color: #e8f2ff;
            }}

            .debug-report {{
                border-left-color: #2e7d32;
                background-color: #eaf7ea;
            }}

            .trace-report {{
                border-left-color: #777777;
                background-color: #eeeeee;
            }}
        </style>
    </head>
    <body>
        <div class="report">
        </div>
    </body>
        <h1>Analysis Report</h1>

        <p><strong>Keyword:<strong> {keyword}</p>
        <p><strong>Levels:</strong> {levels}</p>
        <p><strong>Matches:</strong> {matches}</p>

        <h2>Matching Lines</h2>
        <ul>
            {result_lines}
        </ul>
    </body>
    </html>
    """

    file_data=BytesIO()
    file_data.write(html.encode("utf-8"))
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_report.html",
        mimetype="text/html"
    )

@app.route("/download-analysis-excel")
def download_analysis_excel():

    latest = history[-1] if history else {}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Analysis Results"

    sheet["A1"] = "Level"
    sheet["B1"] = "Message"

    for line in latest.get("results", []):
        line = line.replace("<marks>", "")
        line = line.replace("</marks>", "")

        parts = line.split(" ", 1)

        if len(parts) ==2:
            level = parts[0]
            message = parts[1]
        else:
            level = "UNKNOWN"
            message = line

        sheet.append([level, message])

    file_data = BytesIO()
    workbook.save(file_data)
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/download-analysis-pdf")
def download_analysis_pdf():

    latest = history[-1] if history else {}

    file_data =BytesIO()

    doc = SimpleDocTemplate(
        file_data,
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    styles = getSampleStyleSheet()

    from reportlab.lib.enums import TA_LEFT

    title_style = styles["Title"].clone("LeftTitle")
    title_style.alignment = TA_LEFT

    critical_style = styles["Normal"].clone("CriticalStyle")
    critical_style.textColor = darkred

    error_style = styles["Normal"].clone("ErrorStyle")
    error_style.textColor = red

    warning_style = styles["Normal"].clone("WarningStyle")
    warning_style.textColor = orange

    info_style = styles["Normal"].clone("InfoStyle")
    info_style.textColor = blue

    debug_style = styles["Normal"].clone("DebugStyle")
    debug_style.textColor = green

    trace_style = styles["Normal"].clone("TraceStyle")
    trace_style.textColor = gray
    
    story = []

    story.append(Paragraph("Analysis Report", title_style))
    story.append(Paragraph(f"Keyword: {latest.get('keyword', 'Not set')}", styles["Normal"]))
    story.append(Paragraph(f"Levels: {latest.get('levels', '')}", styles["Normal"]))
    story.append(Paragraph(f"Matches: {latest.get('matches', 0)}", styles["Normal"]))
    story.append(Paragraph("<br/><b>Matching Lines</b>", styles["Heading2"]))

    for line in latest.get("results", []):
        line = line.replace("<mark>", "")
        line = line.replace("</mark>", "")

        if line.startswith("CRITICAL"):
            style = critical_style
        elif line.startswith("ERROR"):
            style = error_style
        elif line.startswith("WARNING"):
            style = warning_style
        elif line.startswith("INFO"):
            style = info_style
        elif line.startswith("DEBUG"):
            style = debug_style
        elif line.startswith("TRACE"):
            style = trace_style
        else:
            style = styles["Normal"]
                                                    
        story.append(Paragraph(line, style))

    doc.build(story)

    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="analysis_report.pdf",
        mimetype="application/pdf"
    )
    

                    
    
if __name__ == "__main__":
    app.run(debug=True)


































