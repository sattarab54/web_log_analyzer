import app as app_module
from app import app
from io import BytesIO
from openpyxl import load_workbook

def test_standard_log_submission():
    client = app.test_client()

    response = client.post(
        "/",
        data={
            "log_text": "ERROR Database connection failed",
            "keyword": "",
            "levels": ["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG", "TRACE"],
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Detected Log Format:" in response.data
    assert b"Standard" in response.data
    assert b"Total matching lines:" in response.data
    assert b"1" in response.data

def test_json_log_submission():
    client = app.test_client()

    response = client.post(
        "/",
        data={
            "log_text": '{"level":"INFO","message":"User logged in"}',
            "keyword": "",
            "levels": ["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG", "TRACE"],
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Detected Log Format:" in response.data
    assert b"JSON" in response.data
    assert b"Total matching lines:" in response.data
    assert b"1" in response.data

def test_apache_log_submission():
    client = app.test_client()

    response = client.post(
        "/",
        data={
            "log_text": '127.0.0.1 - - [26/Aug/2026:10:30:00] "GET /login HTTP/1.1" 404 123',
            "keyword": "",
            "levels": ["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG", "TRACE"],
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Detected Log Format:" in response.data
    assert b"Apache" in response.data
    assert b"Total matching lines:" in response.data
    assert b"1" in response.data

def test_mixed_log_submission():
    client = app.test_client()

    log_text = "\n".join([
        "ERROR Database connection failed",
        '{"level":"INFO","message":"User logged in"}',
        '127.0.0.1 - - [26/Aug/2026:10:30:00] "Get /login HTTP/1.1" 404 123',
    ])

    response = client.post(
        "/",
        data={
            "log_text": log_text,
            "keyword": "",
            "levels": ["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG", "TRACE"],
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Detected Log Format:" in response.data
    assert b"Mixed" in response.data
    assert b"Processed Lines:" in response.data
    assert b"3" in response.data

def test_ignored_line_diagnostics():
    client = app.test_client()

    log_text = "\n".join([
        "ERROR Database connection failed",
        "",
        '{"level":"INFO","message":',
        '127.0.0.1 - - [26/Aug/2026:10:30:00] "GET /login HTTP/1.1"',
        "this is not a recognized log line",
    ])

    response = client.post(
        "/",
        data={
            "log_text": log_text,
            "keyword": "",
            "levels": ["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG", "TRACE"],
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Ignored Lines:" in response.data
    assert b"4" in response.data
    assert b"Blank Lines:" in response.data
    assert b"Invalid JSON Lines:" in response.data
    assert b"Invalid Apache Lines:" in response.data
    assert b"Unknown Lines:" in response.data

def test_processing_rate():
    client = app.test_client()

    log_text = "\n".join([
        "ERROR Database connection failed",
        '{"level":"INFO","message":"User logged in"}',
        "",                
        "this is not a recognized log line",
    ])

    response = client.post(
        "/",
        data={
            "log_text": log_text,
            "keyword": "",
            "levels": ["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG", "TRACE"],
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Total Input Lines:" in response.data
    assert b"4" in response.data
    assert b"Processed Lines:" in response.data
    assert b"2" in response.data
    assert b"Processing Rate:" in response.data
    assert b"50.0%" in response.data

def test_history_keyword_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_search=login"
    )

    assert response.status_code == 200
    assert b"login" in response.data
    assert b"payment" not in response.data

def test_history_level_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_level=ERROR"
    )

    assert response.status_code == 200
    assert b"login" in response.data
    assert b"payment" not in response.data

def test_history_from_date_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-25 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_from=2026-08-22"
    )

    assert response.status_code == 200
    assert b"payment" in response.data
    assert b"login" not in response.data
    
def test_history_to_date_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-25 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_to=2026-08-22"
    )

    assert response.status_code == 200
    assert b"login" in response.data
    assert b"payment" not in response.data
    
def test_history_combined_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "login",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-23 10:00:00",
        },        
        {
            "keyword": "payment",
            "levels": "ERROR",
            "matches": 4,
            "searched_at": "2026-08-24 10:00:00",
        },        
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?"
        "history_search=login&"
        "history_level=ERROR&"
        "history_from=2026-08-19&"
        "history_to=2026-08-21"
    )

    assert response.status_code == 200
    assert b"login" in response.data
    assert b"payment" not in response.data
    assert b"2026-08-23 10:00:00" not in response.data

def test_history_no_filters(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-25 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get("/filter-history")

    assert response.status_code == 200
    assert b"login" in response.data
    assert b"payment" in response.data

def test_history_sort_newest(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "older",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "newer",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-25 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_sort=newest"
    )

    assert response.status_code == 200
    page =  response.data.decode()
    assert page.find("newer") != -1
    assert page.find("older") != -1
    assert page.find("newer") < page.find("older")

def test_history_sort_oldest(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "older",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "newer",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-25 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_sort=oldest"
    )

    assert response.status_code == 200
    page =  response.data.decode()
    assert page.find("newer") != -1
    assert page.find("older") != -1
    assert page.find("older") < page.find("newer")

def test_history_sort_keyword_asc(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "zebra",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "apple",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-25 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_sort=keyword_asc"
    )

    assert response.status_code == 200
    page =  response.data.decode()
    assert page.find("apple") != -1
    assert page.find("zebra") != -1
    assert page.find("apple") < page.find("zebra")

def test_history_sort_keyword_desc(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "apple",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "zebra",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-25 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_sort=keyword_desc"
    )

    assert response.status_code == 200
    page =  response.data.decode()
    assert page.find("zebra") != -1
    assert page.find("apple") != -1    
    assert page.find("zebra") < page.find("apple")

def test_history_pagination_page_1(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": f"item{i}",
            "levels": "INFO",
            "matches": i,
            "searched_at": f"2026-08-{i:02d} 10:00:00",
        }
        for i in range(1, 13)        
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_sort=oldest&page=1"
    )

    assert response.status_code == 200

    page =  response.data.decode()
    assert "<td>item2</td>" in page
    assert "<td>item10</td>" in page
    assert "<td>item11</td>" not in page
    assert "<td>item12</td>" not in page
    
def test_history_pagination_page_2(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": f"item{i}",
            "levels": "INFO",
            "matches": i,
            "searched_at": f"2026-08-{i:02d} 10:00:00",
        }
        for i in range(1, 13)        
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_sort=oldest&page=2"
    )

    assert response.status_code == 200

    page =  response.data.decode()
    assert "<td>item11</td>" in page
    assert "<td>item12</td>" in page
    assert "<td>item10</td>" not in page
    
def test_history_pagination_out_of_range(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": f"item{i}",
            "levels": "INFO",
            "matches": i,
            "searched_at": f"2026-08-{i:02d} 10:00:00",
        }
        for i in range(1, 13)        
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/filter-history?history_sort=oldest&page=99"
    )

    assert response.status_code == 200

    page =  response.data.decode()   

    assert "<td>item11</td>" not in page
    assert "<td>item12</td>" not in page

def test_download_results_csv():
    client = app.test_client()

    response = client.get("/download-csv")

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    assert "analysis_results.csv" in response.headers["Content-Disposition"]

def test_download_history_csv():
    client = app.test_client()

    response = client.get("/download-history-csv")

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    
def test_download_filtered_history_csv(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-history-csv?history_search=login"
    )

    assert response.status_code == 200
    assert b"login" in response.data
    assert b"payment" not in response.data
    
def test_download_filtered_history_csv_level_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-history-csv?history_level=ERROR"
    )

    assert response.status_code == 200
    assert b"login" in response.data
    assert b"payment" not in response.data

def test_download_filtered_history_csv_date_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "old",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-10 10:00:00",
        },
        {
            "keyword": "middle",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "new",
            "levels": "WARNING",
            "matches": 3,
            "searched_at": "2026-08-30 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-history-csv?"
        "history_from=2026-08-15&"
        "history_to=2026-08-25"
    )

    assert response.status_code == 200
    assert b"middle" in response.data
    assert b"old" not in response.data
    assert b"new" not in response.data

def test_download_filtered_history_csv_combined_filters(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
        },
        {
            "keyword": "login",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-20 11:00:00",
        },
        {
            "keyword": "payment",
            "levels": "ERROR",
            "matches": 4,
            "searched_at": "2026-08-20 12:00:00",
        },
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 5,
            "searched_at": "2026-08-30 10:00:00",
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-history-csv?"
        "history_search=login&"
        "history_level=ERROR&"
        "history_from=2026-08-15&"
        "history_to=2026-08-25"
    )

    assert response.status_code == 200
    assert b"login,ERROR,3" in response.data
    assert b"login,WARNING,2" not in response.data
    assert b"payment,ERROR,4" not in response.data
    assert b"login,ERROR,5" not in response.data

def test_download_history_excel():
    client = app.test_client()

    response = client.get("/download-history-excel")

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "history.xlsx" in response.headers["Content-Disposition"]

def test_download_filtered_history_excel():
    client = app.test_client()

    response = client.get("/download-filtered-history-excel")

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
def test_filtered_history_excel_keyword_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_search=login"
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    keywords = [
        sheet.cell(row=row, column=1).value
        for row in range(2, sheet.max_row + 1)
    ]

    assert "login" in keywords
    assert "payment" not in keywords

def test_filtered_history_excel_level_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_level=ERROR"
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    levels = [
        sheet.cell(row=row, column=2).value
        for row in range(2, sheet.max_row + 1)
    ]

    assert "ERROR" in levels
    assert "WARNING" not in levels

def test_filtered_history_excel_date_filter(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "old",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-10 10:00:00",
            "results": [],
        },
        {
            "keyword": "middle",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "new",
            "levels": "WARNING",
            "matches": 3,
            "searched_at": "2026-08-30 10:00:00",
            "results": [],
        }
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?"
        "history_from=2026-08-15&"
        "history_to=2026-08-25"
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    keywords = [
        sheet.cell(row=row, column=1).value
        for row in range(2, sheet.max_row + 1)
    ]

    assert "middle" in keywords
    assert "oldt" not in keywords
    assert "new" not in keywords

def test_filtered_history_excel_combined_filters(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "login",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-20 11:00:00",
            "results": [],
        },
        {
            "keyword": "payment",
            "levels": "ERROR",
            "matches": 4,
            "searched_at": "2026-08-20 12:00:00",
            "results": [],
        },
        {
            "keyword": "login",
            "levels": "ERROR",
            "matches": 5,
            "searched_at": "2026-08-30 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?"
        "history_search=login&"
        "history_level=ERROR&"
        "history_from=2026-08-15&"
        "history_to=2026-08-25"
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    rows = [
        (
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
        )
        for row in range(2, sheet.max_row + 1)
    ]

    assert ("login", "ERROR", 3) in rows
    assert ("login", "WARNING", 2) not in rows
    assert ("payment", "ERROR", 4) not  in rows
    assert ("login", "ERROR", 5) not in rows
    
def test_filtered_history_excel_sort_newest(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "older",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "newer",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-25 11:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=newest"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    keywords = [
        sheet.cell(row=row, column=1).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert keywords == ["newer", "older"]
    
def test_filtered_history_excel_sort_oldest(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "older",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "newer",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-25 11:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=oldest"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    keywords = [
        sheet.cell(row=row, column=1).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert keywords == ["older", "newer"]

def test_filtered_history_excel_sort_keyword_az(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "payment",
            "levels": "WARNING",
            "matches": 2,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "error",
            "levels": "ERROR",
            "matches": 3,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },
        {
            "keyword": "login",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-22 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=keyword_asc"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    keywords = [
        sheet.cell(row=row, column=1).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert keywords == ["error", "login", "payment"]

def test_filtered_history_excel_sort_keyword_desc(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "apple",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "zebra",
            "levels": "ERROR",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },
        {
            "keyword": "login",
            "levels": "WARNING",
            "matches": 3,
            "searched_at": "2026-08-22 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=keyword_desc"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    keywords = [
        sheet.cell(row=row, column=1).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert keywords == ["zebra", "login", "apple"]

def test_filtered_history_excel_sort_levels_high(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "info-item",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "critical-item",
            "levels": "CRITICAL",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },
        {
            "keyword": "warning-item",
            "levels": "WARNING",
            "matches": 3,
            "searched_at": "2026-08-22 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=levels_high"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    levels = [
        sheet.cell(row=row, column=2).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert levels == ["CRITICAL", "WARNING", "INFO"]

def test_filtered_history_excel_sort_levels_low(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "critical-item",
            "levels": "CRITICAL",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "info-item",
            "levels": "INFO",
            "matches": 2,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },        
        {
            "keyword": "trace-item",
            "levels": "TRACE",
            "matches": 3,
            "searched_at": "2026-08-22 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=levels_low"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    levels = [
        sheet.cell(row=row, column=2).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert levels == ["TRACE", "INFO", "CRITICAL"]
    
def test_filtered_history_excel_sort_matches_high(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "low",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "high",
            "levels": "ERROR",
            "matches": 5,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },        
        {
            "keyword": "middle",
            "levels": "WARNING",
            "matches": 3,
            "searched_at": "2026-08-22 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=matches_high"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    matches = [
        sheet.cell(row=row, column=3).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert matches == [5, 3, 1]

def test_filtered_history_excel_sort_matches_low(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "high",
            "levels": "ERROR",
            "matches": 5,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "low",
            "levels": "INFO",
            "matches": 1,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },        
        {
            "keyword": "middle",
            "levels": "WARNING",
            "matches": 3,
            "searched_at": "2026-08-22 10:00:00",
            "results": [],
        },
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_sort=matches_low"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    matches = [
        sheet.cell(row=row, column=3).value            
        for row in range(2, sheet.max_row + 1)
    ]

    assert matches == [1, 3, 5]

def test_filtered_history_excel_empty_result(monkeypatch):
    client = app.test_client()

    test_history = [
        {
            "keyword": "login",
            "levels": "INFO",
            "matches": 3,
            "searched_at": "2026-08-20 10:00:00",
            "results": [],
        },
        {
            "keyword": "payment",
            "levels": "ERROR",
            "matches": 5,
            "searched_at": "2026-08-21 10:00:00",
            "results": [],
        },        
    ]

    monkeypatch.setattr(app_module, "history", test_history)

    response = client.get(
        "/download-filtered-history-excel?history_search=not-found"        
    )

    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["Filtered History"]

    assert sheet.max_row == 1

    headers = [
        sheet.cell(row=1, column=column).value            
        for column in range(1, 5)
    ]

    assert headers == [
        "Keyword",
        "Levels",
        "Matches",
        "Searched At",
    ]



    

































